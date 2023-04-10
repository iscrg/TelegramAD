from telethon import TelegramClient
from openpyxl import load_workbook
import socks

inviters = []
wb = load_workbook('data/inviters.xlsx')
sheet = wb.get_sheet_by_name('Лист1')
rows = sheet.max_row
cols = sheet.max_column

for i in range(2, rows + 1):
    inviters.append([])
    for j in range(1, cols + 1):
        cell = sheet.cell(row = i, column = j)
        inviters[i-2].append(str(cell.value))

for inviter in inviters:
    api_id = int(inviter[0])
    api_hash = inviter[1]
    phone_number = inviter[2]
    try:
        with TelegramClient(phone_number, api_id, api_hash) as client:
            print('success!')
    except:
        print('error!')