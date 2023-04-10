from telethon import TelegramClient
from telethon.errors.rpcerrorlist import PeerFloodError
from telethon.tl.functions.account import GetAuthorizationsRequest
from openpyxl import load_workbook
import subprocess
import sys
from time import sleep

#Read ToDoList sheet
wb = load_workbook('__todolist_invite.xlsx')
sheet = wb.get_sheet_by_name('Лист1')
rows = sheet.max_row
cols = sheet.max_column

bots = [0] * (rows-1)

#Parallel launch of bots
for i in range(2, rows + 1):
    string = ''
    for j in range(1, cols + 1):
        cell = sheet.cell(row = i, column = j)
        string = string + str(cell.value) + '|'

    with open('bots/' + 'tasks/' + str(i-1) + '.txt', 'wb') as file:
        file.write(string)
    
    file_name = 'bots/' + str(i-1) + '.py'
    bots[i-2] = subprocess.Popen([sys.executable, file_name])
    bots[i-2].wait()
    sleep(2)

for i in range(0, rows-1):
    bots[i].wait()