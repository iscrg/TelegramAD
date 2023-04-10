from telethon import TelegramClient
from openpyxl import load_workbook
import socks
import asyncio
import python_socks

proxy_http = (socks.SOCKS5, '188.143.169.27', 40046, False, 'iparchitect_22157_23_01_23', 'FDRyZk5K2n94ihQBNH')
wb = load_workbook('__todolist.xlsx')
sheet = wb.get_sheet_by_name('Лист1')
rows = sheet.max_row

report = ''

for i in range(1, rows):
    with open('bots/report/' + str(i) + '.txt') as file:
        report += file.readline() + '\n'

async def do(proxy):
        client = TelegramClient(
        session='79639471107',
        api_id=15601172,
        api_hash='30f1741b7586eff23b65b2734bdfe896',
        proxy=proxy
        )
        await client.connect()
        await client.send_message('79833134667', report)

if __name__ == '__main__':
    asyncio.set_event_loop(asyncio.SelectorEventLoop())
    loop = asyncio.get_event_loop()
    loop.run_until_complete(do(proxy_http))