from telethon.sync import TelegramClient
from telethon.tl.types import InputPeerUser
from telethon.errors.rpcerrorlist import PeerFloodError, UserPrivacyRestrictedError
from telethon.tl.functions.channels import InviteToChannelRequest
import traceback
import time
import random
import socks
from openpyxl import load_workbook


task_num = 1        

#Read the task.
with open(f'bots/tasks/{str(task_num)}.txt') as file:
    data = file.readline().split('|')
mode = int(data[3])

#Read proxy.
with open('bots/data/proxy.txt') as file:
    proxy_socks5 = file.readline().split(':')
proxy_socks5 = (socks.SOCKS5, 
                proxy_socks5[0], 
                int(proxy_socks5[1]), 
                False, 
                proxy_socks5[3], 
                proxy_socks5[4])

#Read inviting accounts.
inviters = []
wb = load_workbook('bots/data/inviters.xlsx')
sheet = wb.get_sheet_by_name('Лист1')
rows = sheet.max_row
cols = sheet.max_column

for i in range(2, rows + 1):
    inviters.append([])
    for j in range(1, cols + 1):
        cell = sheet.cell(row = i, column = j)
        inviters[i-2].append(str(cell.value))

inviter_counter = int(data[6])

api_id = inviters[inviter_counter-1][0]
api_hash = inviters[inviter_counter-1][1]
phone = inviters[inviter_counter-1][2]

#Read users.
users = []
wb = load_workbook('bots/data/' + data[1])
sheet = wb.get_sheet_by_name('Лист1')
rows = sheet.max_row
cols = sheet.max_column

for i in range(1, rows + 1):
    users.append([])
    for j in range(1, cols + 1):
        cell = sheet.cell(row = i, column = j)
        users[i-2].append(str(cell.value))
users = users[int(data[4])-1 : int(data[5])]

#Connect to account.
try:
    client = TelegramClient(phone, 
                            api_id, 
                            api_hash, 
                            proxy=proxy_socks5)
    client.connect()
    target_group_entity = client.get_entity(data[2])
except:
    with open(f'bots/report/{str(task_num)}.txt', 'r+') as file:
        file.write(f'task-{str(task_num)}, name: {data[0]}, LOGIN ERROR PLS CHECK BOT {str(inviter_counter)}')
    exit()

n = 0
counter = 0
flood = []
errors = 0

for user in users:
    counter += 1

    #Change bot after 30 invites.
    if counter == 30:
        inviter_counter += 1
        client.close()
        api_id = inviters[inviter_counter-1][0]
        api_hash = inviters[inviter_counter-1][1]
        phone = inviters[inviter_counter-1][2]
        client = TelegramClient(phone, api_id, api_hash, proxy=proxy_socks5)
        try:
            client.connect()
        except:
            with open('bots/report/' + str(task_num) + '.txt', 'r+') as file:
                file.write('task-' + str(task_num) + ', ' + 'name: ' + data[0] + ', ' + 'LOGIN ERROR PLS CHECK BOT ' + str(inviter_counter))
            break
        counter = 0
    
    if 1 == 1:
        time.sleep(1)

        #Invite user.
        try:
            print ("Adding {}".format(user[0]))
            if mode == 1:
                user_to_add = client.get_input_entity(user[0])
            elif mode == 2:
                user_to_add = InputPeerUser(user[0], user[1])
            client(InviteToChannelRequest(target_group_entity,[user_to_add]))
            time.sleep(random.randrange(10, 30))

        except PeerFloodError:
            print(f'[!] Task-{str(task_num)}: Getting Flood Error from telegram. Script is stopping now. Please try again after some time.')
            flood.append(inviter_counter)
            continue

        except UserPrivacyRestrictedError:
            print(f"[!] Task-{str(task_num)} The user's privacy settings do not allow you to do this. Skipping.")
            errors += 1
            continue

        except:
            traceback.print_exc()
            print(f'[!] Task-{str(task_num)} Unexpected Error')
            errors += 1
            continue

#Writing report.
with open(f'bots/report/{str(task_num)}.txt', 'wb') as file:            
    if flood == []:
        file.write(f'Task-{str(task_num)}, Name: {data[0]}, Errors: {str(errors)}')
    else:
        file.write(f'Task-{str(task_num)}, Name: {data[0]}, Errors: {str(errors)} [!] FLOOD ERROR! PLS CHECK BOT {str(flood)}')