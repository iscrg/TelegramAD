from openpyxl import load_workbook
import subprocess
import sys
from time import sleep
from math import ceil

#Read ToDoList.
wb = load_workbook('__todolist_spam.xlsx')
sheet = wb.get_sheet_by_name('Лист1')
rows = sheet.max_row
cols = sheet.max_column

bots = [0] * (rows-1)
bot_number = 0

for i in range(2, rows + 1):
    #Writing tasks to bots.
    string = ''
    for j in range(1, 5):
        cell = sheet.cell(row = i, column = j)
        string = string + str(cell.value) + '|'
    start_bot = int(sheet.cell(row = i, column = 7).value)
    finish_bot = int(sheet.cell(row = i, column = 8).value)
    start_user= int(sheet.cell(row = i, column = 5).value)
    finish_user = int(sheet.cell(row = i, column = 6).value)
    for g in range(start_bot, finish_bot):
        string1 = string + str(start_user)
        start_user += 45
        if start_user < finish_user:
            string1 = string1 + '|' + str(start_user)
        else:
            string1 = string1 + '|' + str(finish_user)
        with open('bots/' + 'tasks/' + str(g) + '.txt', 'wb') as file:
            file.write(string1)
    
    #Parallel launch of bots.
    file_name = 'bots/' + str(i-1) + '.py'
    bots[i-2] = subprocess.Popen([sys.executable, file_name])
    bots[i-2].wait()
    sleep(2)

for i in range(0, rows-1):
    bots[i].wait()