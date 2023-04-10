from telethon.sync import TelegramClient
from telethon import functions
from telethon.sync import TelegramClient
from telethon.tl.types import InputPeerUser
from telethon.errors.rpcerrorlist import PeerFloodError
import traceback
from time import sleep
from random import randint
import socks
from openpyxl import load_workbook

bot = 1

#Read the task.
with open('bots/tasks/' + str(bot) + '.txt') as file:
    data = file.readline().split('|')
mode = int(data[3])

#Read proxy.
with open('bots/data/proxy.txt') as file:
    proxy_socks5 = file.readline().split(':')
print(proxy_socks5)
proxy_socks5 = (socks.SOCKS5, proxy_socks5[0], int(proxy_socks5[1]), False, proxy_socks5[2], proxy_socks5[3])

#Read bot data.
inviters = []
wb = load_workbook('bots/data/inviters.xlsx')
sheet = wb.get_sheet_by_name('Лист1')
rows = sheet.max_row
cols = sheet.max_column

for i in range(2, rows + 1):
    inviters.append([])
    for j in range(1, cols + 1):
        cell = sheet.cell(row = i, column = j)
        inviters[i-2].append(str(cell.value))

api_id = int(inviters[bot-1][0])
api_hash = inviters[bot-1][1]
phone = inviters[bot-1][2]

#Read users to invite.
users = []
wb = load_workbook('bots/data/' + data[1])
sheet = wb.get_sheet_by_name('Лист1')
rows = sheet.max_row
cols = sheet.max_column

for i in range(1, rows + 1):
    users.append([])
    for j in range(1, cols + 1):
        cell = sheet.cell(row = i, column = j)
        users[i-2].append(str(cell.value))
users = users[int(data[4])-1:int(data[5])]

#Connect to account
try:
    with TelegramClient(phone, api_id, api_hash, proxy=proxy_socks5) as client:
        #Sending messages
        errors = 0
        client.get_dialogs()
        for user in users:
            try:
                if mode == 1:
                    user_to_send = client.get_input_entity(user[0])
                elif mode == 2:
                    user_to_send = InputPeerUser(user[0], user[1])

                client(functions.messages.ForwardMessagesRequest(
                from_peer=1861197001,
                id=[int(data[2])],
                to_peer=user_to_send,
                ))
            except PeerFloodError:
                print('bot-' + str(bot) + ' ' +"[!] Getting Flood Error from telegram. \n[!] Script is stopping now. \n[!] Please try again after some time.")
                with open('bots/report/' + str(bot) + '.txt', 'w') as file:
                    file.write('bot:' + bot + ' name: ' + data[0] + ', ' + 'errors: ' + str(errors+ (45-users.index(user))) + 'FLOOD ERROR! PLS CHECK BOT')
                exit()
            except:
                traceback.print_exc()
                print('bot-' + str(bot) + ' ' + "[!] Unexpected Error")
                errors += 1
                continue
            sleep(randint(15, 30))

        with open('bots/report/' + str(bot) + '.txt', 'wb') as file:
            file.write('bot:' + str(bot) + ' name: ' + str(data[0]) + ', ' + 'errors: ' + str(errors))

except:
    with open('bots/report/' + str(bot) + '.txt', 'wb') as file:
        file.write('name: ' + data[0] + ', ' + 'LOGIN ERROR PLS CHECK BOT ' + str(bot))
    exit() 


